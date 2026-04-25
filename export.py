import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side, numbers)
from openpyxl.utils import get_column_letter
from datetime import datetime

BLUE = "FF0070C0"
DARK_BLUE = "FF002060"
LIGHT_BLUE = "FFDCE6F1"
GREEN = "FF00B050"
RED = "FFFF0000"
YELLOW = "FFFFFF00"
HEADER_FILL = PatternFill("solid", start_color=DARK_BLUE)
ALT_FILL = PatternFill("solid", start_color=LIGHT_BLUE)
WHITE_FILL = PatternFill("solid", start_color="FFFFFFFF")

thin = Side(style="thin", color="FFD3D3D3")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

#helper cosmetic cell function
def _hdr(cell, text, width=None):
    cell.value = text
    cell.font = Font(bold=True, color="FFFFFFFF", name="Arial", size=10)
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER

#helper to format decimal values as percentages
def _fmt_pct(cell):
    cell.number_format = "0.00%;(0.00%);-"

#helper to format to specified decimal values with a default of two
def _fmt_num(cell, decimals=2):
    fmt = f"0.{'0'*decimals};(0.{'0'*decimals});-"
    cell.number_format = fmt

#styler that writes value, applies alternating row fill, right-aligns, adds border, then applies percentage/number formatting
def _body(cell, value, pct=False, decimals=2, alt=False):
    cell.value = value
    cell.font = Font(name="Arial", size=10)
    cell.fill = ALT_FILL if alt else WHITE_FILL
    cell.alignment = Alignment(horizontal="right")
    cell.border = BORDER
    if pct:
        _fmt_pct(cell)
    else:
        _fmt_num(cell, decimals)

#helper function to differentiate ticker row headings
def _ticker_cell(cell, ticker, alt=False):
    cell.value = ticker
    cell.font = Font(bold=True, name="Arial", size=10, color=DARK_BLUE)
    cell.fill = ALT_FILL if alt else WHITE_FILL
    cell.alignment = Alignment(horizontal="left")
    cell.border = BORDER

"""
This report expects:
-General metrics
-A correlation matrix
-prices
-The users list of stock tickers
-The initial distribution of stock weights in portfolio
-The temporary xlsx's path
-The list of stocks that passed and failed the screening
-Drifted allocation data (end_vals, start_vals, drifted_weights, total_end)
"""
def export_report(metrics: pd.DataFrame, corr: pd.DataFrame, prices: pd.DataFrame,
                  tickers: list, weights: dict, output_path: str, screened: pd.DataFrame = None,
                  end_vals: dict = None, start_vals: dict = None, drifted_weights: dict = None, total_end: float = None):
    wb = Workbook()

    #Sheet 1: Summary Metrics
    ws = wb.active
    ws.title = "Summary Metrics"
    ws.sheet_view.showGridLines = False

    #Title
    ws.merge_cells("A1:G1")
    title_cell = ws["A1"]
    title_cell.value = f"Stock Screener & Portfolio Report  |  Generated {datetime.today().strftime('%B %d, %Y')}"
    title_cell.font = Font(bold=True, size=14, color="FFFFFFFF", name="Arial")
    title_cell.fill = HEADER_FILL
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28
    #spacer
    ws.row_dimensions[2].height = 6

    #Column headers
    headers = ["Ticker", "Ann. Return", "Volatility", "Sharpe Ratio", "Beta", "Max Drawdown", "Screening"]
    cols = ["A", "B", "C", "D", "E", "F", "G"]
    widths = [12, 14, 13, 14, 10, 16, 14]
    for col, hdr, w in zip(cols, headers, widths):
        _hdr(ws[f"{col}3"], hdr)
        ws.column_dimensions[col].width = w
    ws.row_dimensions[3].height = 22

    #Data rows
    if screened is not None:
        screened_tickers = set(screened.index.tolist())
    else:
        screened_tickers = set(metrics.index.tolist())

    for i, (ticker, row) in enumerate(metrics.iterrows(), start=4):
        #boolean to represent if you are on an even row
        alt = i % 2 == 0
        _ticker_cell(ws[f"A{i}"], ticker, alt)
        _body(ws[f"B{i}"], row["Ann. Return"], pct=True, alt=alt)
        _body(ws[f"C{i}"], row["Volatility"], pct=True, alt=alt)
        _body(ws[f"D{i}"], row["Sharpe Ratio"], decimals=2, alt=alt)
        _body(ws[f"E{i}"], row["Beta"], decimals=2, alt=alt)
        _body(ws[f"F{i}"], row["Max Drawdown"], pct=True, alt=alt)
        passes = ticker in screened_tickers
        #assigned passed screen status
        ws[f"G{i}"].value = "✓ Pass" if passes else "✗ Fail"
        ws[f"G{i}"].font = Font(name="Arial", size=10, bold=True, color=("FF00B050" if passes else "FFFF0000"))
        ws[f"G{i}"].fill = ALT_FILL if alt else WHITE_FILL
        ws[f"G{i}"].alignment = Alignment(horizontal="center")
        ws[f"G{i}"].border = BORDER

    #Sheet 2: Correlation Matrix 
    ws2 = wb.create_sheet("Correlation Matrix")
    ws2.sheet_view.showGridLines = False
    ws2.merge_cells(f"A1:{get_column_letter(len(tickers)+1)}1")
    t2 = ws2["A1"]
    t2.value = "Return Correlation Matrix"
    t2.font = Font(bold=True, size=13, color="FFFFFFFF", name="Arial")
    t2.fill = HEADER_FILL
    t2.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 26

    ws2["A2"].value = ""
    ws2["A2"].fill = HEADER_FILL
    #copy the tickers into the columns and rows
    for j, t in enumerate(corr.columns, start=2):
        col_letter = get_column_letter(j)
        ws2.column_dimensions[col_letter].width = 11
        _hdr(ws2[f"{col_letter}2"], t)
    ws2.column_dimensions["A"].width = 11

    for i, row_ticker in enumerate(corr.index, start=3):
        _hdr(ws2[f"A{i}"], row_ticker)
        alt = i % 2 == 0
        for j, col_ticker in enumerate(corr.columns, start=2):
            val = corr.loc[row_ticker, col_ticker]
            cell = ws2.cell(row=i, column=j)
            #populate the excel matrix values with values from the correlation matrix
            cell.value = round(val, 3)
            cell.font = Font(name="Arial", size=10)
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="center")
            cell.number_format = "0.000"
            #Assign color gradient. green=high corr, white=low
            if row_ticker == col_ticker:
                cell.fill = PatternFill("solid", start_color="FF002060")
                cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFFFF")
            elif val >= 0.7:
                cell.fill = PatternFill("solid", start_color="FFC6EFCE")
            elif val >= 0.4:
                cell.fill = PatternFill("solid", start_color="FFEBF5EB")
            elif val < 0:
                cell.fill = PatternFill("solid", start_color="FFFFC7CE")
            else:
                cell.fill = WHITE_FILL

    #Sheet 3: Portfolio
    ws3 = wb.create_sheet("Portfolio")
    ws3.sheet_view.showGridLines = False
    for col, w in zip(["A","B","C","D","E","F"], [14,16,20,16,18,18]):
        ws3.column_dimensions[col].width = w

    #Section 1: Target Allocation
    ws3.merge_cells("A1:C1")
    t3 = ws3["A1"]
    t3.value = "Target Allocation"
    t3.font = Font(bold=True, size=13, color="FFFFFFFF", name="Arial")
    t3.fill = HEADER_FILL
    t3.alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 26

    for col, hdr in zip(["A", "B", "C"], ["Ticker", "Weight", "Allocated Value ($100k)"]):
        _hdr(ws3[f"{col}2"], hdr)

    total_port = 100_000
    #populate columns A, B, and C with Ticker, weight, and allocated value per 100k respectively
    for i, (t, w) in enumerate(weights.items(), start=3):
        alt = i % 2 == 0
        _ticker_cell(ws3[f"A{i}"], t, alt)
        _body(ws3[f"B{i}"], w, pct=True, alt=alt)
        _body(ws3[f"C{i}"], total_port * w, decimals=0, alt=alt)
        ws3[f"C{i}"].number_format = '$#,##0;($#,##0);-'

    #index of the last data row in the target allocation table
    last = 2 + len(weights)
    
    ws3[f"A{last+1}"].value = "Total"
    ws3[f"A{last+1}"].font = Font(bold=True, name="Arial")
    ws3[f"B{last+1}"] = f"=SUM(B3:B{last})"
    ws3[f"B{last+1}"].number_format = "0.00%;(0.00%);-"
    ws3[f"B{last+1}"].font = Font(bold=True, name="Arial")
    ws3[f"C{last+1}"] = f"=SUM(C3:C{last})"
    ws3[f"C{last+1}"].number_format = '$#,##0;($#,##0);-'
    ws3[f"C{last+1}"].font = Font(bold=True, name="Arial")

    #Section 2: Drifted Allocation
    if drifted_weights and end_vals and start_vals and total_end:
        #leave two blank rows as spacers
        drift_start_row = last + 4  

        ws3.merge_cells(f"A{drift_start_row}:D{drift_start_row}")
        drift_title = ws3[f"A{drift_start_row}"]
        drift_title.value = f"Current Drifted Allocation  |  Total Portfolio Value: ${total_end:,.0f}"
        drift_title.font = Font(bold=True, size=13, color="FFFFFFFF", name="Arial")
        drift_title.fill = HEADER_FILL
        drift_title.alignment = Alignment(horizontal="center", vertical="center")
        ws3.row_dimensions[drift_start_row].height = 26

        hdr_row = drift_start_row + 1
        #format column headers
        for col, hdr in zip(["A","B","C","D"], ["Ticker","Drifted Weight","Current Value","Gain / Loss"]):
            _hdr(ws3[f"{col}{hdr_row}"], hdr)

        #populate columns A, B, C, D with Ticker, drifted weight, Current Value and Profit respectively
        for j, (t, dw) in enumerate(drifted_weights.items(), start=hdr_row + 1):
            alt = j % 2 == 0
            _ticker_cell(ws3[f"A{j}"], t, alt)
            _body(ws3[f"B{j}"], dw, pct=True, alt=alt)

            ev = end_vals.get(t, 0)
            sv = start_vals.get(t, 0)
            gl = ev - sv

            ws3[f"C{j}"].value = ev
            ws3[f"C{j}"].font = Font(name="Arial", size=10)
            ws3[f"C{j}"].fill = ALT_FILL if alt else WHITE_FILL
            ws3[f"C{j}"].alignment = Alignment(horizontal="right")
            ws3[f"C{j}"].border = BORDER
            ws3[f"C{j}"].number_format = '$#,##0;($#,##0);-'

            ws3[f"D{j}"].value = gl
            ws3[f"D{j}"].font = Font(name="Arial", size=10, color="FF00B050" if gl >= 0 else "FFFF0000")
            ws3[f"D{j}"].fill = ALT_FILL if alt else WHITE_FILL
            ws3[f"D{j}"].alignment = Alignment(horizontal="right")
            ws3[f"D{j}"].border = BORDER
            ws3[f"D{j}"].number_format = '$#,##0_);[Red]($#,##0)'

        drift_last = hdr_row + len(drifted_weights)
        ws3[f"A{drift_last+1}"].value = "Total"
        ws3[f"A{drift_last+1}"].font = Font(bold=True, name="Arial")
        ws3[f"C{drift_last+1}"] = f"=SUM(C{hdr_row+1}:C{drift_last})"
        ws3[f"C{drift_last+1}"].number_format = '$#,##0;($#,##0);-'
        ws3[f"C{drift_last+1}"].font = Font(bold=True, name="Arial")
        ws3[f"D{drift_last+1}"] = f"=SUM(D{hdr_row+1}:D{drift_last})"
        ws3[f"D{drift_last+1}"].number_format = '$#,##0_);[Red]($#,##0)'
        ws3[f"D{drift_last+1}"].font = Font(bold=True, name="Arial")

    #Sheet 4: Price History
    ws4 = wb.create_sheet("Price History")
    ws4.sheet_view.showGridLines = False
    ws4.merge_cells(f"A1:{get_column_letter(len(tickers)+1)}1")
    t4 = ws4["A1"]
    t4.value = "Normalized Price History (Base = 100)"
    t4.font = Font(bold=True, size=13, color="FFFFFFFF", name="Arial")
    t4.fill = HEADER_FILL
    t4.alignment = Alignment(horizontal="center", vertical="center")
    ws4.row_dimensions[1].height = 26
    ws4.column_dimensions["A"].width = 14
    _hdr(ws4["A2"], "Date")

    #create a column for every ticker
    valid_tickers = [t for t in tickers if t in prices.columns]
    for j, t in enumerate(valid_tickers, start=2):
        col_l = get_column_letter(j)
        ws4.column_dimensions[col_l].width = 12
        _hdr(ws4[f"{col_l}2"], t)

    #normalize prices to 100 and drop blank entries
    norm = prices[valid_tickers].dropna()
    norm = (norm / norm.iloc[0]) * 100
    for i, (date, row_data) in enumerate(norm.iterrows(), start=3):
        #give A entry in the row a value of the date
        ws4[f"A{i}"] = date.strftime("%Y-%m-%d") if hasattr(date, "strftime") else str(date)
        ws4[f"A{i}"].font = Font(name="Arial", size=9)
        ws4[f"A{i}"].border = BORDER
        alt = i % 2 == 0
        for j, t in enumerate(valid_tickers, start=2):
            cell = ws4.cell(row=i, column=j)
            cell.value = round(row_data[t], 2)
            cell.font = Font(name="Arial", size=9)
            cell.border = BORDER
            cell.number_format = "0.00"
            cell.fill = ALT_FILL if alt else WHITE_FILL

    wb.save(output_path)
    return output_path